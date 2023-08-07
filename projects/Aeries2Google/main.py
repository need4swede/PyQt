## IMPORTS
if 'Imports':

    if 'Standard':
        import os, sys, csv, warnings, re, subprocess, platform
        from datetime import datetime

    if 'Libraries':
        try:
            from n4s import fs # pip install n4s
        except ImportError:
            print("\n'n4s' library not installed! Please install it via 'pip install n4s'\n")
            sys.exit()
        from openpyxl import load_workbook
        warnings.filterwarnings("ignore")

    if 'PyQt6':
        from PyQt6.QtWidgets import (QApplication, QWidget, QLabel, QHBoxLayout, QVBoxLayout,
                                    QLineEdit, QPushButton, QTextEdit, QFileDialog, QMainWindow)
        from PyQt6.QtGui import QPalette, QColor
        from PyQt6.QtCore import Qt, QDir, QTimer

## SETTINGS
if 'Settings':

    ## APP INFO
    APP_NAME = "Aeries2Google"

    ## APP VERSION
    APP_VERSION = 1.0

    ## APP DEVS
    APP_DEVS = "need4swede"

    ## APP YEAR
    APP_YEAR = datetime.today().strftime("%Y")

    ## DOMAIN
    DOMAIN = "need4swede.net"

    ## PASSWORD
    PASSWORD = "need4swede"

    ## ORG UNIT
    ORG_UNIT = "/Students"

## GLOBAL VARIABLES
if 'Global Variables':

    ## OS VERSION
    OS_VERSION = float(fs.system('info')[1])

    ## USER DIR
    USER = f"{QDir.homePath()}"

    ## SCREEN WIDTH
    SCREEN_WIDTH = ''

    ## SCREEN HEIGHT
    SCREEN_HEIGHT = ''

## TERMINAL WINDOW
class Terminal(QWidget):

    ## CONSTRUCTOR
    def __init__(self):

        ## INITIALIZE PARENT CLASSES
        super().__init__()

        ## SET WINDOW SIZE
        self.setFixedSize(425, 170)

        ## CREATE TEXT EDIT WIDGET
        self.text_edit = QTextEdit()
        self.text_edit.setReadOnly(True)

        ## CREATE LAYOUT
        layout = QVBoxLayout()
        layout.addWidget(self.text_edit)

        ## SET LAYOUT
        self.setLayout(layout)

        ## SET WINDOW TITLE
        self.setWindowTitle(APP_NAME)

        ## INITIAL STATEMENT FLAG
        self.first_statement = True

    ## METHOD TO PRINT OUTPUTS
    def print(self, statement, spacing=1):

        ## SET NUMBER OF NEWLINES
        newlines = '\n' * spacing

        ## IF NOT THE FIRST STATEMENT, ADD TWO LINE SPACES
        if not self.first_statement:
            statement = newlines + statement

        ## APPEND STATEMENT TO THE TEXT EDIT
        self.text_edit.append(statement)

        ## UPDATE FIRST STATEMENT FLAG
        self.first_statement = False

    ## METHOD TO CLEAR TERMINAL
    def clear(self):

        ## CLEAR THE TEXT EDIT
        self.text_edit.clear()

        ## RESET FIRST STATEMENT FLAG
        self.first_statement = True

    ## OVERRIDE THE CLOSE EVENT
    def closeEvent(self, event):
        QApplication.quit()

## PAUSES THE TERMINAL
class PauseTimer:

    ## INITIALIZE CLASS
    def __init__(self, milliseconds, callback):

        ## CREATE TIMER
        self.timer = QTimer()

        ## SET TIMER INTERVAL
        self.timer.setInterval(milliseconds)

        ## CONNECT CALLBACK TO TIMER TIMEOUT SIGNAL
        self.timer.timeout.connect(callback)

        ## CONNECT CALLBACK TO STOP TIMER
        self.timer.timeout.connect(self.stop)

    ## START TIMER
    def start(self):
        self.timer.start()

    ## STOP TIMER
    def stop(self):
        self.timer.stop()

## SET UI THEMES
class ThemeHandler:
    ## SET THEME METHOD
    def set_theme(self, theme):
        ## CREATE PALETTE
        palette = QPalette()

        ## SET DARK THEME
        if theme.lower() == 'dark':
            ## THEME SETTINGS
            palette.setColor(QPalette.ColorRole.Window, QColor(53, 53, 53))
            palette.setColor(QPalette.ColorRole.WindowText, Qt.GlobalColor.white)
            palette.setColor(QPalette.ColorRole.Base, QColor(25, 25, 25))
            palette.setColor(QPalette.ColorRole.AlternateBase, QColor(53, 53, 53))
            palette.setColor(QPalette.ColorRole.ToolTipBase, Qt.GlobalColor.black)
            palette.setColor(QPalette.ColorRole.ToolTipText, Qt.GlobalColor.white)
            palette.setColor(QPalette.ColorRole.Text, Qt.GlobalColor.white)
            palette.setColor(QPalette.ColorRole.Button, QColor(53, 53, 53))
            palette.setColor(QPalette.ColorRole.ButtonText, Qt.GlobalColor.white)
            palette.setColor(QPalette.ColorRole.BrightText, Qt.GlobalColor.red)
            palette.setColor(QPalette.ColorRole.Link, QColor(42, 130, 218))
            palette.setColor(QPalette.ColorRole.Highlight, Qt.GlobalColor.transparent)
            palette.setColor(QPalette.ColorRole.HighlightedText, Qt.GlobalColor.black)

        ## APPLY THEME
        QApplication.setPalette(palette)

## FILE BROWSER
class FileBrowser(QWidget, ThemeHandler):

    ## CONSTRUCTOR
    def __init__(self):

        ## INITIALIZE PARENT CLASSES
        super().__init__()

        self.setStyleSheet("""
            QWidget {
                background-color: #333333;
                color: #ffffff;
            }
            QPushButton {
                background-color: #555555;
                color: #ffffff;
                border: none;
                padding: 5px;
                border-radius: 3px;
                width:60px;
            }
            QPushButton:hover {
                background-color: #666666;
            }
            QLineEdit {
                background-color: #444444;
                color: #ffffff;
                height: 24px;
            }
        """)

        ## INITIALIZE TERMINAL WINDOW
        self.terminal_window = Terminal()
        self.terminal_window.print('Processing...')

        ## INITIALIZE SPREADSHEET HANDLER
        self.spreadsheet = SpreadsheetHandler()

        ## INITIALIZE PAUSE TIMER
        self.pause = None
        self.finish = None

        ## CREATE LAYOUT
        layout = QHBoxLayout()
        layout.setContentsMargins(15, 5, 15, 0)
        layout.setSpacing(10)

        ## CREATE FILE PATH EDIT
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText('Path to Aeries Export...')
        self.file_path_edit.setStyleSheet("QLineEdit { border: 1px solid #555555; border-radius: 5px; }")  # Added rounded edges
        self.file_path_edit.textChanged.connect(self.validate_file)  # Connect to new method
        layout.addWidget(self.file_path_edit)

        ## CREATE BROWSE BUTTON
        self.browse_button = QPushButton('Browse')
        self.browse_button.clicked.connect(self.browse_file)
        self.browse_button.setToolTip('Browse and select a file')
        layout.addWidget(self.browse_button)

        ## CREATE OPEN BUTTON
        self.open_button = QPushButton('Open')
        self.open_button.setEnabled(False)
        self.open_button.clicked.connect(self.open_file)
        self.open_button.setToolTip('Open the selected file')
        layout.addWidget(self.open_button)

        ## SET THEME
        self.set_theme('dark')

        ## SET LAYOUT
        self.setLayout(layout)

        ## CREATE APP DIRECTORY
        self.create_appdir('parent')
        self.create_appdir('child')

    ## CREATE APP DIRECTORY
    def create_appdir(self, dir):

        ## Get user's home directory
        home_dir = os.path.expanduser("~")

        ## Append 'Documents' to the home directory
        documents_dir = os.path.join(home_dir, 'Documents')

        ## Append 'Aeries2Google' to the Documents directory
        self.app_dir = os.path.join(documents_dir, 'Aeries2Google')

        ## MAIN APP DIR
        if dir == 'parent':

            ## CREATE PARENT DIR
            fs.path_exists(self.app_dir, True)

        ## SUBDIRS WITH DATES
        if dir == 'child':

            ## TODAY'S DATE
            today = datetime.now().strftime('%Y-%m-%d')

            ## CREATE CHILD DIR
            child = os.path.join(self.app_dir, today)
            fs.path_exists(child, True)

            ## RETURN CHILD PATH
            return child

    ## VALIDATE FILE SELECTION
    def validate_file(self):

        ## INITIALIZE VALIDATION FLAG
        is_valid = False

        ## GET FILENAME
        file = os.path.basename(self.file_path_edit.text())

        ## VALIDATE
        is_aeries_export = file.startswith('PrintQueryToExcel')
        is_xlsx = file.endswith('.xlsx')
        if is_aeries_export and is_xlsx:
            is_valid = True

        ## ENABLE / DISABLE 'OPEN' BUTTON
        self.open_button.setEnabled(is_valid)

        # Check if the file name starts with "PrintQueryToExcel" and extract the date
        self.xlsx_file_export_date = None
        if is_valid:
            date_str = file.split('_')[-3]  # Extract date string after the last underscore
            if len(date_str) == 8 and date_str.isdigit():  # Check if it's a valid date format
                try:
                    year = int(date_str[:4])
                    month = int(date_str[4:6])
                    day = int(date_str[6:8])
                    month_name = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"][month - 1]
                    self.xlsx_file_export_date = f"{month_name} {day}, {year}"
                except ValueError as e:
                    print(f"'{e}' :: on line {sys.exc_info()[-1].tb_lineno}")

    ## BROWSE FILE METHOD
    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(directory=USER)
        self.file_path_edit.setText(file_path)

    ## OPEN FILE
    def open_file(self):

        ## PATH OF SELECTED FILE
        file_path = self.file_path_edit.text()

        ## TERMINAL OUTPUT
        self.terminal_window.show()

        ## CREATE PAUSE INSTANCE
        self.pause = PauseTimer(1500, lambda: self.process_file(file_path))

        ## START THE PAUSE
        self.pause.start()

    ## READ CONTENTS OF IMPORT
    def process_file(self, xlsx_file):

        csv_file = os.path.join(self.create_appdir('child'), f'users_{datetime.now().strftime("%H_%M_%S")}.csv')

        ## TERMINAL OUTPUT
        self.terminal_window.print("Detected Aeries Export")
        self.terminal_window.print(f"Export Date: {self.xlsx_file_export_date}", 0)
        self.terminal_window.print(f"Entries Found: {self.spreadsheet.count_xlsx_rows(xlsx_file)}", 0)

        ## COPY DATA FROM FILE TO USERS.CSV
        self.spreadsheet.create_csv(csv_file)
        self.spreadsheet.copy_names(xlsx_file, csv_file)
        self.spreadsheet.create_emails(xlsx_file, csv_file)
        self.spreadsheet.create_passwords(csv_file)
        self.spreadsheet.create_org_units(csv_file)

        ## COMPLETE PROCESSING
        self.terminal_window.print('File Processing Complete')
        self.finish = PauseTimer(2000, lambda: self.finished_file(csv_file))
        self.finish.start()

    ## PROCESS COMPLETE
    def finished_file(self, csv_file):

        ## Depending on the OS, use different commands
        if platform.system() == 'Windows':  # for Windows
            subprocess.Popen(f'explorer /select,"{csv_file}"')
        elif platform.system() == 'Darwin':  # for MacOS
            subprocess.Popen(['open', '-R', csv_file])
        else:  # for Linux
            subprocess.Popen(['xdg-open', csv_file])

## HANDLES SPREADSHEET OPERATIONS
class SpreadsheetHandler:

    ## CONSTRUCTOR
    def __init__(self):

        ## INITIALIZE PARENT CLASSES
        super().__init__()

    ## READ THE NUMBER OF ROWS IN A GIVEN .XLSX FILE
    def count_xlsx_rows(self, xlsx_file):

        ## LOAD THE .XLSX WORKBOOK
        workbook = load_workbook(xlsx_file)
        sheet = workbook.active

        ## COUNT THE NUMBER OF ROWS EXCLUDING THE HEADER ROW
        row_count = sum(1 for row in sheet.iter_rows(min_row=2) if any(cell.value for cell in row))

        ## RETURN ROW COUNT
        return row_count

    ## GENERATE 'USERS' CSV FILE
    def create_csv(self, csv_file):

        ## GOOGLE IMPORT HEADERS
        headers = [
            "First Name [Required]",
            "Last Name [Required]",
            "Email Address [Required]",
            "Password [Required]",
            "Password Hash Function [UPLOAD ONLY]",
            "Org Unit Path [Required]",
            "New Primary Email [UPLOAD ONLY]",
            "Recovery Email",
            "Home Secondary Email",
            "Work Secondary Email",
            "Recovery Phone [MUST BE IN THE E.164 FORMAT]",
            "Work Phone",
            "Home Phone",
            "Mobile Phone",
            "Work Address",
            "Home Address",
            "Employee ID",
            "Employee Type",
            "Manager Email",
            "Department",
            "Cost Center",
            "Building ID",
            "Floor Name",
            "Floor Section",
            "New Status [UPLOAD ONLY]",
            "Advanced Protection Program enrollment"
        ]

        ## SAVE FILE
        with open(csv_file, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(headers)

    ## COPY DATA FROM IMPORT TO 'USERS' CSV FILE
    def copy_names(self, xlsx_file, csv_file):

        ## LOAD THE .XLSX WORKBOOK
        workbook = load_workbook(xlsx_file)
        sheet = workbook.active

        ## FUNCTION TO REPLACE SPACES AND CONSECUTIVE DASHES WITH A SINGLE DASH
        def replace_invalid_characters(value):
            if value:
                value = re.sub(' +', ' ', value)  # Replace consecutive spaces with a single space
                value = value.replace(' - ', '-')  # Replace space-dash-space with a single dash
                value = value.replace(' ', '-')     # Replace remaining spaces with a dash
                value = re.sub('-+', '-', value)    # Replace consecutive dashes with a single dash
                value = value.replace('’', '') if value else value
            return value

        ## READ THE "FIRST NAME" COLUMN, REPLACE SPACES WITH DASHES, AND KEEP ONLY UTF-8 COMPLIANT CHARACTERS
        first_names = [replace_invalid_characters(cell.value).encode('utf-8', 'ignore').decode('utf-8') for cell in sheet['B'] if cell.row > 1]
        first_names = [name.title() for name in first_names]  # Capitalize names

        ## READ THE "LAST NAME" COLUMN, REPLACE SPACES WITH DASHES, AND KEEP ONLY UTF-8 COMPLIANT CHARACTERS
        last_names = [replace_invalid_characters(cell.value).encode('utf-8', 'ignore').decode('utf-8') for cell in sheet['C'] if cell.row > 1]
        last_names = [name.title() for name in last_names]  # Capitalize names

        ## OPEN THE 'USERS.CSV' FILE AND UPDATE THE "FIRST NAME [REQUIRED]" AND "LAST NAME [REQUIRED]" COLUMNS
        with open(csv_file, 'r') as file:
            reader = csv.reader(file)
            rows = list(reader)

        ## ENSURE THERE ARE ENOUGH ROWS IN THE CSV TO ACCOMMODATE THE FIRST NAMES AND LAST NAMES
        while len(rows) < max(len(first_names), len(last_names)) + 1:
            rows.append([''] * len(rows[0]))

        ## UPDATE THE "FIRST NAME [REQUIRED]" COLUMN WITH THE FIRST_NAMES LIST
        for idx, name in enumerate(first_names):
            rows[idx + 1][0] = name

        ## UPDATE THE "LAST NAME [REQUIRED]" COLUMN WITH THE LAST_NAMES LIST
        for idx, name in enumerate(last_names):
            rows[idx + 1][1] = name

        ## WRITE THE UPDATED ROWS BACK TO 'USERS.CSV'
        with open(csv_file, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerows(rows)

    ## FUNCTION TO GENERATE EMAILS FROM FIRST NAME AND STUDENT ID
    def create_emails(self, xlsx_file, csv_file):

        ## LOAD THE .XLSX WORKBOOK
        workbook = load_workbook(xlsx_file)
        sheet = workbook.active

        ## READ FIRST NAMES
        with open(csv_file, 'r') as file:
            reader = csv.reader(file)
            rows = list(reader)
            first_names = [row[0].split('-')[0] for row in rows[1:]]

        ## READ THE "STUDENT ID" COLUMN
        student_ids = [cell.value for cell in sheet['E'] if cell.row > 1]

        ## GENERATE EMAILS
        emails = [f"{first_name.lower()}.{student_id}@{DOMAIN}" for first_name, student_id in zip(first_names, student_ids)]

        ## OPEN THE 'USERS.CSV' FILE AND UPDATE THE "EMAIL ADDRESS [REQUIRED]" COLUMN
        with open(csv_file, 'r') as file:
            reader = csv.reader(file)
            rows = list(reader)

        ## ENSURE THERE ARE ENOUGH ROWS IN THE CSV TO ACCOMMODATE THE EMAILS
        while len(rows) < len(emails) + 1:
            rows.append([''] * len(rows[0]))

        ## UPDATE THE "EMAIL ADDRESS [REQUIRED]" COLUMN WITH THE EMAILS LIST
        for idx, email in enumerate(emails):
            rows[idx + 1][2] = email  # Column index 2 corresponds to "Email Address [Required]"

        ## WRITE THE UPDATED ROWS BACK TO 'USERS.CSV'
        with open(csv_file, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerows(rows)

    ## SETS PASSWORD FOR EVERY USER
    def create_passwords(self, csv_file):

        ## OPEN THE 'USERS.CSV' FILE AND UPDATE THE "PASSWORD [REQUIRED]" COLUMN
        with open(csv_file, 'r') as file:
            reader = csv.reader(file)
            rows = list(reader)

        ## UPDATE THE "PASSWORD [REQUIRED]" COLUMN WITH THE PASSWORD
        for row in rows[1:]:  # Exclude the header row
            row[3] = PASSWORD  # Column index 3 corresponds to "Password [Required]"

        ## WRITE THE UPDATED ROWS BACK TO 'USERS.CSV'
        with open(csv_file, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerows(rows)

    ## SETS ORG UNITS FOR EVERY USER
    def create_org_units(self, csv_file):

        ## OPEN THE 'USERS.CSV' FILE AND UPDATE THE "ORG UNIT PATH [REQUIRED]" COLUMN
        with open(csv_file, 'r') as file:
            reader = csv.reader(file)
            rows = list(reader)

        ## UPDATE THE "ORG UNIT PATH [REQUIRED]" COLUMN WITH THE ORG_UNIT
        for row in rows[1:]:  # Exclude the header row
            row[5] = ORG_UNIT  # Column index 5 corresponds to "Org Unit Path [Required]"

        ## WRITE THE UPDATED ROWS BACK TO 'USERS.CSV'
        with open(csv_file, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerows(rows)

## APP CLASS
class Application(QMainWindow, ThemeHandler):

    ## CONSTRUCTOR
    def __init__(self):

        ## INITIALIZE PARENT CLASSES
        super().__init__()

        ## ENABLE DRAG AND DROP
        self.setAcceptDrops(True)

        ## INITIALIZE UI
        self.initUI()

    ## HANDLES DRAGGING
    def dragEnterEvent(self, event):
        # When a drag event enters the widget, we need to determine whether we should accept it.
        # In this case, we only accept it if it contains file URLs.
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.file_browser.setVisible(False)  # Hide the file browser
            self.drop_label.setVisible(True)  # Show the drop label

    ## HANDLES DROPPING
    def dropEvent(self, event):
        # When a file is dropped on the widget, get the first URL.
        # We take the first URL only because we only handle one file.
        file_url = event.mimeData().urls()[0]

        # Assuming the FileBrowser is the central widget of your application,
        # you can access the file_path_edit QLineEdit like this:
        self.file_browser.file_path_edit.setText(file_url.toLocalFile())

        # Validate the file and, if valid, programmatically click the 'open' button
        self.file_browser.validate_file()
        if self.file_browser.open_button.isEnabled():
            self.file_browser.open_button.click()

    ## HANDLES DRAG LEAVING
    def dragLeaveEvent(self, event):
        self.file_browser.setVisible(True)  # Show the file browser
        self.drop_label.setVisible(False)  # Hide the drop label

    ## INIT UI METHOD
    def initUI(self):
        ## SET WINDOW TITLE
        self.setWindowTitle("Aeries2Google")

        ## CREATE MAIN LAYOUT
        main_layout = QVBoxLayout()

        ## CREATE FILE BROWSER
        self.file_browser = FileBrowser()

        ## CREATE DROP LABEL
        self.drop_label = QLabel("Drop XLSX File")
        self.drop_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.drop_label.setStyleSheet("font-size: 24px;")
        self.drop_label.setVisible(False)  # Initially hidden

        ## CREATE MAIN LAYOUT
        main_layout = QVBoxLayout()

        ## ADD WIDGETS TO LAYOUT
        main_layout.addWidget(self.file_browser)
        main_layout.addWidget(self.drop_label)

        ## CREATE MAIN WIDGET AND SET LAYOUT
        main_widget = QWidget()
        main_widget.setLayout(main_layout)

        ## SET CENTRAL WIDGET
        self.setCentralWidget(main_widget)

        ## CREATE FOOTER LABEL
        footer_label = QLabel(f'Version {APP_VERSION} - Developed by {APP_DEVS}, {APP_YEAR}')
        footer_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        footer_label.setStyleSheet("color: #888888; font-size: 8pt;") # Styling for footer text
        main_layout.addWidget(footer_label)

        ## CREATE MAIN WIDGET AND SET LAYOUT
        main_widget = QWidget()
        main_widget.setLayout(main_layout)

        ## SET CENTRAL WIDGET
        self.setCentralWidget(main_widget)

        ## SET FIXED HEIGHT
        self.setFixedSize(380, 80)

        ## SHOW WINDOW
        self.show()

## MAIN
if __name__ == '__main__':
    ## CREATE APPLICATION
    app = QApplication(sys.argv)

    ## CREATE APP INSTANCE
    Aeries2Google = Application()

    ## EXECUTE APPLICATION
    sys.exit(app.exec())